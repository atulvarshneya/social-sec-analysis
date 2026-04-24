from src.config import config
from src.validators import validate_data_not_empty
from src.logger import logger, log_report_generated, log_file_error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side, GradientFill, Alignment
import os

def reports(all_results: dict):
    """Generate Excel reports from analysis results.
    
    Args:
        all_results: Dictionary of scenario results to report on
        
    Raises:
        ValueError: If results data is empty
        Exception: If Excel file creation or writing fails
    """
    try:
        validate_data_not_empty(all_results, "Analysis results")
    except ValueError as e:
        logger.error(str(e))
        raise
    
    try:
        # Create output directory if it doesn't exist
        output_dir = os.path.dirname(config.COMBINED_BENEFITS_FILE)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"Created output directory: {output_dir}")
        
        wb = Workbook()

        # Collect all scenario names in sorted order
        scenario_names = sorted(list(all_results.keys()))
        # Collect all years across all scenarios to ensure we cover the full range in the report
        all_years = set()
        for scenario_name, scenario_data in all_results.items():
            for year in scenario_data.keys():
                all_years.add(year)

        ## REPORT 1: Combined report across all scenarios showing monthly benefits and cumulative totals
        ws = wb.active
        ws.title = "Combined Scenario Benefits"

        # Initialize cumulative tracker for each scenario
        scenario_cumulative = {scname: 0.0 for scname in scenario_names}  # Initialize cumulative tracker for each scenario

        # Add a header row in Excel sheet with columns for each scenario's monthly benefit and cumulative total
        header_row = ["Year-Month"]
        for scname in scenario_names:
            header_row.append(f"{scname} Benefit")
            header_row.append(f"{scname} Cumulative")
        ws.append(header_row)  # Sub-header row in Excel
        for cell in ws[1]:  # Apply bold font to header row
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="00FFFF")
            cell.alignment = Alignment(horizontal='general', vertical='center', wrap_text=True, shrink_to_fit=True)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            ws.column_dimensions[column].width = 20

        # Add rows for each month and scenario's paid benefit, while calculating cumulative totals
        for year in sorted(all_years):
            for month in range(1, 13):
                row = []
                row.append(f"{year:04d}-{month:02d}")
                for scname in scenario_names:
                    benefit = all_results[scname].get(year, {}).get(month, {}).get('paid_benefit', 'N/A')
                    if benefit != 'N/A':
                        scenario_cumulative[scname] += benefit
                    row.append(benefit)
                    row.append(scenario_cumulative[scname])
                ws.append(row)  # Add the row to the Excel sheet
                for cell in ws[ws.max_row]:  # Apply number formatting to the new row
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'  # Format numbers with two decimal places and comma as thousand separator
                    else:
                        cell.number_format = '@'  # Treat non-numeric values as text
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # REPORT 2: Each scenario Monthly details with eligible benefit, income, phase, and paid benefit
        for scname in scenario_names:
            ws = wb.create_sheet(title=f"{scname} Monthly Details")
            ws.title = f"{scname} Monthly Details"

            scenario_results = all_results[scname]

            header_row = ["Year-Month", "Eligible Benefit", "Monthly Income", "Phase", "Income Limit", "Paid Benefit", "Cumulative"]
            ws.append(header_row)  # Sub-header row in Excel
            for cell in ws[1]:  # Apply bold font to header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="00FFFF")
                cell.alignment = Alignment(horizontal='general', vertical='center', wrap_text=True, shrink_to_fit=True)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                ws.column_dimensions[column].width = 20

            cumulative_paid_benefit = 0.0
            for year in sorted(all_years):
                for month in range(1, 13):
                    cumulative_paid_benefit += scenario_results.get(year, {}).get(month, {}).get('paid_benefit', 0.0)
                    row = []
                    row.append(f"{year}-{month:02d}")
                    row.append(scenario_results.get(year, {}).get(month, {}).get('eligible_benefit', 'N/A'))
                    row.append(scenario_results.get(year, {}).get(month, {}).get('monthly_income', 'N/A'))
                    row.append(scenario_results.get(year, {}).get(month, {}).get('phase', 'N/A'))
                    row.append(scenario_results.get(year, {}).get(month, {}).get('income_limit', 'N/A'))
                    row.append(scenario_results.get(year, {}).get(month, {}).get('paid_benefit', 'N/A'))
                    row.append(cumulative_paid_benefit)
                    ws.append(row)
                    for cell in ws[ws.max_row]:  # Apply number formatting to the new row
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'  # Format numbers with two decimal places and comma as thousand separator
                        else:
                            cell.number_format = '@'  # Treat non-numeric values as text
                            cell.alignment = Alignment(horizontal="center", vertical="center")

        # REPORT 3: Each scenario Annual summary with total eligible benefit, total income, and total paid benefit
        for scname in scenario_names:
            ws = wb.create_sheet(title=f"{scname} Annual Summary")
            ws.title = f"{scname} Annual Summary"

            scenario_results = all_results[scname]

            header_row = ["Year", "Total Eligible Benefit", "Total Income", "Total Paid Benefit"]
            ws.append(header_row)  # Sub-header row in Excel
            for cell in ws[1]:  # Apply bold font to header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="00FFFF")
                cell.alignment = Alignment(horizontal='general', vertical='center', wrap_text=True, shrink_to_fit=True)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                ws.column_dimensions[column].width = 20

            for year in sorted(all_years):
                total_eligible_benefit = sum(details['eligible_benefit'] for details in scenario_results.get(year, {}).values())
                total_income = sum(details['monthly_income'] for details in scenario_results.get(year, {}).values())
                total_paid_benefit = sum(details['paid_benefit'] for details in scenario_results.get(year, {}).values())
                row = [f"{year}", total_eligible_benefit, total_income, total_paid_benefit]
                ws.append(row)
                for cell in ws[ws.max_row]:  # Apply number formatting to the new row
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'  # Format numbers with two decimal places and comma as thousand separator
                    else:
                        cell.number_format = '@'  # Treat non-numeric values as text
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Save the reports Excel file
        wb.save(config.COMBINED_BENEFITS_FILE)
        log_report_generated(config.COMBINED_BENEFITS_FILE, len(wb.sheetnames))
        
    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}")
        raise